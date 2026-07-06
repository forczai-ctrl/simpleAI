import io
import csv
from typing import Any

try:
    import pypdf
except ImportError:
    pypdf = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


def parse_pdf(file_bytes: bytes) -> str:
    """Extract plain text from a PDF file."""
    if pypdf is None:
        raise ImportError("pypdf is not installed in the environment.")
    
    text_content = []
    try:
        pdf_file = io.BytesIO(file_bytes)
        reader = pypdf.PdfReader(pdf_file)
        for i, page in enumerate(reader.pages):
            page_text = page.extract_text()
            if page_text:
                text_content.append(page_text)
        return "\n--- Page Break ---\n".join(text_content)
    except Exception as e:
        raise ValueError(f"Error parsing PDF file: {str(e)}")


def parse_excel(file_bytes: bytes) -> str:
    """Convert Excel spreadsheet rows into an LLM-readable text format."""
    if openpyxl is None:
        raise ImportError("openpyxl is not installed in the environment.")
    
    output = []
    try:
        excel_file = io.BytesIO(file_bytes)
        workbook = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            output.append(f"Sheet: {sheet_name}")
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                # Filter out completely empty rows
                if not any(cell is not None for cell in row):
                    continue
                row_str = " | ".join(str(cell) if cell is not None else "" for cell in row)
                output.append(f"Row {row_idx}: {row_str}")
        return "\n".join(output)
    except Exception as e:
        raise ValueError(f"Error parsing Excel file: {str(e)}")


def parse_csv(file_bytes: bytes) -> str:
    """Convert CSV file rows into text format."""
    try:
        text_data = file_bytes.decode("utf-8", errors="ignore")
        reader = csv.reader(io.StringIO(text_data))
        output = []
        for i, row in enumerate(reader, 1):
            if not row:
                continue
            row_str = " | ".join(row)
            output.append(f"Row {i}: {row_str}")
        return "\n".join(output)
    except Exception as e:
        raise ValueError(f"Error parsing CSV file: {str(e)}")
